from openpyxl import load_workbook


def _norm_header(s: str) -> str:
    """
    Нормализация заголовка для устойчивого сопоставления.

    Убирает регистр, пробелы и точки. Это позволяет читать xlsx с
    вариациями: 'ед. изм', 'едизм', 'Ед.Изм' — все попадают в одну
    каноническую форму 'едизм'.
    """
    return s.lower().replace(".", "").replace(" ", "")


class ExcelReader:
    def __init__(self, path: str, sheet_name: str = None):
        self.path = path
        self.sheet_name = sheet_name

    def read_rows(self):
        wb = load_workbook(self.path, data_only=True)

        ws = wb[self.sheet_name] if self.sheet_name else wb.active

        # Карта: нормализованное имя заголовка -> номер колонки.
        # Пустые ячейки в шапке игнорируются.
        headers: dict[str, int] = {}
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value is None:
                continue
            headers[_norm_header(str(cell.value))] = idx

        # Канонические имена колонок (как они написаны в "эталонном" шаблоне).
        # Сопоставляются с реальными заголовками через _norm_header.
        required = [
            "Участок линии",
            "Единица промышленного оборудования",
            "Комплектующие",
            "Количество комплектующих",
            "ед. изм",
            "Себестоимость комплектующих, с НДС",
            "тип",
        ]

        for col in required:
            if _norm_header(col) not in headers:
                raise ValueError(f"Нет колонки '{col}' в Excel")

        # Хелпер: достать номер колонки по каноническому имени.
        def col_idx(name: str) -> int:
            return headers[_norm_header(name)]

        for row_idx in range(2, ws.max_row + 1):
            yield {
                "row_number": row_idx,
                "area":              ws.cell(row_idx, col_idx("Участок линии")).value,
                "equipment_name":    ws.cell(row_idx, col_idx("Единица промышленного оборудования")).value,
                "component_name":    ws.cell(row_idx, col_idx("Комплектующие")).value,
                "qty_per_equipment": ws.cell(row_idx, col_idx("Количество комплектующих")).value,
                "unit":              ws.cell(row_idx, col_idx("ед. изм")).value,
                "component_type":    ws.cell(row_idx, col_idx("тип")).value,
                "cost_with_vat":     ws.cell(row_idx, col_idx("Себестоимость комплектующих, с НДС")).value,
            }
