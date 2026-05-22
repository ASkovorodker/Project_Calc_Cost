"""
Генератор result.xlsx из шаблона.

Колонки находятся по именам шапки (с нормализацией: lower, без пробелов
и точек) — благодаря этому шаблон можно безопасно перекраивать
(переставлять колонки, добавлять новые, чуть менять написание), не правя
код генератора. Если в шаблоне нет какой-то обязательной колонки —
ExcelGenerator падает с понятной ошибкой.

Поведение:
  * Открываем копию шаблона.
  * Читаем шапку листа "Базовый", строим карту (нормализованное имя → col_idx).
  * Для каждого dict из data ищем колонки по каноническим именам
    и записываем значения.
  * Сохраняем как output_path.
"""
from pathlib import Path
from typing import List, Dict
from datetime import datetime
import shutil

from openpyxl import load_workbook


def _norm_header(s: str) -> str:
    """Нормализация заголовка: lower, без пробелов и точек.

    Тот же подход, что и в parser/excel/reader.py. Снимает класс
    проблем с написанием колонок ('ед. изм' vs 'ед.изм.' vs 'Ед Изм').
    """
    return s.lower().replace(".", "").replace(" ", "")


# Маппинг: ключ в data dict → каноническое имя колонки в шаблоне.
# Если в шаблоне реальное название слегка отличается — _norm_header
# приведёт к общему виду.
COLUMN_MAP = (
    ("line_section",       "Участок линии"),
    ("status",             "Статус подбора из БД"),
    ("equipment",          "Единица промышленного оборудования"),
    ("component_name",     "Комплектующие"),
    ("equipment_quantity", "Количество единиц"),
    ("component_quantity", "Количество комплектующих"),
    ("unit",               "ед. изм"),
    ("cost_with_vat",      "Себестоимость комплектующих, с НДС"),
    ("type",               "тип"),
)


class ExcelGeneratorError(Exception):
    pass


class ExcelGenerator:
    def __init__(self, template_path: str):
        self.template_path = Path(template_path)

        if not self.template_path.exists():
            raise ExcelGeneratorError(
                f"Шаблон не найден: {self.template_path}"
            )

    def generate(
        self,
        data: List[Dict],
        output_path: str | None = None,
    ) -> str:
        """
        Генерирует новый Excel на основе template.xlsx
        """
        output_file = self._prepare_output_file(output_path)

        wb = load_workbook(output_file)
        ws = wb["Базовый"]

        # Карта: нормализованное имя заголовка → номер колонки.
        # Пустые ячейки в шапке игнорируем.
        headers: dict[str, int] = {}
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value is None:
                continue
            headers[_norm_header(str(cell.value))] = idx

        # Проверяем, что все нужные колонки есть в шаблоне.
        missing = []
        col_idx: dict[str, int] = {}
        for data_key, canonical in COLUMN_MAP:
            idx = headers.get(_norm_header(canonical))
            if idx is None:
                missing.append(canonical)
            else:
                col_idx[data_key] = idx
        if missing:
            raise ExcelGeneratorError(
                "В шаблоне нет обязательных колонок: " + ", ".join(missing)
            )

        # Пишем данные.
        current_row = 2  # после шапки
        for row in data:
            for data_key, _canonical in COLUMN_MAP:
                value = row.get(data_key)
                ws.cell(
                    row=current_row,
                    column=col_idx[data_key],
                    value=value,
                )
            current_row += 1

        wb.save(output_file)
        return str(output_file)

    # ----------------------------

    def _prepare_output_file(self, output_path: str | None) -> Path:
        """
        Создаёт копию template.xlsx
        """
        if output_path:
            output_file = Path(output_path)
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = Path(f"generated_{timestamp}.xlsx")

        output_file.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy(self.template_path, output_file)

        return output_file
